# -*- coding: utf-8 -*-
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# ==============================================================================
# CONFIGURATION
# ==============================================================================
CONFIG = {
    "INPUT_FILES": {
        "SciDeBERTa-Pipeline": "./eval/pipeline_eval.json",
        "Gemini-Zero-Shot": "./eval/gemini_zero_shot_v2.json",
        "OpenAI-Zero-Shot": "./eval/openai_v2_zero_shot.json",
        "DeepSeek-Zero-Shot": "./eval/deepseek_v2_zero_shot.json"
    },
    "OUTPUT_EXCEL": "./eval/human_annotation_sample.xlsx",
    "TOTAL_SAMPLE_SIZE": 400, 
    "MIN_PER_PREDICATE": 5,   # Aims for ~20 globally per predicate across the 4 sources
    "NUM_RESEARCHERS": 5,
    "RANDOM_SEED": 42
}

def apply_bold_to_entities(text, sub, obj):
    """Finds the exact entities in the abstract and applies standard Excel bolding."""
    text = str(text) if pd.notna(text) else ""
    sub = str(sub) if pd.notna(sub) else ""
    obj = str(obj) if pd.notna(obj) else ""
    
    entities = sorted(list(set([e for e in [sub, obj] if e])), key=len, reverse=True)
    if not entities:
        return text
    
    escaped_entities = [re.escape(e) for e in entities]
    pattern = re.compile(f"({'|'.join(escaped_entities)})", re.IGNORECASE)
    
    parts = pattern.split(text)
    rich_text_elements = []
    bold_font = InlineFont(b=True) 
    
    for part in parts:
        if any(part.lower() == e.lower() for e in entities):
            rich_text_elements.append(TextBlock(bold_font, part))
        elif part:
            rich_text_elements.append(part)
            
    return CellRichText(*rich_text_elements)

def load_and_sample():
    print(f"\n--- Starting Human Evaluation Sampler ({CONFIG['TOTAL_SAMPLE_SIZE']} Triples, Balanced & Deduplicated) ---")
    
    all_samples = []
    valid_sources = []

    for source_name, filepath in CONFIG["INPUT_FILES"].items():
        if os.path.exists(filepath):
            valid_sources.append((source_name, filepath))
        else:
            print(f"⚠️ Warning: Could not find {filepath}. Skipping.")

    if not valid_sources:
        print("❌ Error: No input files found.")
        return

    samples_per_source = CONFIG["TOTAL_SAMPLE_SIZE"] // len(valid_sources)
    remainder = CONFIG["TOTAL_SAMPLE_SIZE"] % len(valid_sources)

    # 1. Stratified Sampling with Abstract Diversity
    for i, (source_name, filepath) in enumerate(valid_sources):
        df = pd.read_json(filepath)
        df['source'] = source_name 
        
        # Drop internal identical triples first
        df = df.drop_duplicates(subset=['paper_id', 'subject', 'predicate', 'object'])
        
        target_size = samples_per_source + (1 if i < remainder else 0)
        
        # Calculate rarity to prioritize finding unique abstracts for rare predicates
        pred_counts = df['predicate'].value_counts().sort_values()
        
        selected_indices = []
        seen_papers = set()

        # PHASE A: Guarantee minimums for each predicate (hunting rarest first)
        for pred in pred_counts.index:
            # Only look at triples where we haven't used the abstract yet
            available_for_pred = df[(df['predicate'] == pred) & (~df['paper_id'].isin(seen_papers))]
            
            # Don't exceed the overall target size while balancing
            space_left = target_size - len(selected_indices)
            if space_left <= 0:
                break
                
            take_n = min(CONFIG["MIN_PER_PREDICATE"], len(available_for_pred), space_left)
            
            if take_n > 0:
                sampled = available_for_pred.sample(n=take_n, random_state=CONFIG["RANDOM_SEED"])
                selected_indices.extend(sampled.index)
                seen_papers.update(sampled['paper_id'].tolist())

        # PHASE B: Fill the rest of the quota randomly
        remaining_needed = target_size - len(selected_indices)
        if remaining_needed > 0:
            available_remaining = df[(~df.index.isin(selected_indices)) & (~df['paper_id'].isin(seen_papers))]
            
            if len(available_remaining) >= remaining_needed:
                fill_sample = available_remaining.sample(n=remaining_needed, random_state=CONFIG["RANDOM_SEED"])
            else:
                # Absolute fallback if we run out of unique abstracts
                print(f"   ⚠️ Low unique abstracts for {source_name}. Allowing duplicate abstracts to reach {target_size}.")
                available_fallback = df[~df.index.isin(selected_indices)]
                take_fallback = min(remaining_needed, len(available_fallback))
                fill_sample = available_fallback.sample(n=take_fallback, random_state=CONFIG["RANDOM_SEED"])
                
            selected_indices.extend(fill_sample.index)

        sampled_df = df.loc[selected_indices]
        all_samples.append(sampled_df)
        print(f"   ✅ Sampled {len(sampled_df)} balanced triples from {source_name}")

    final_df = pd.concat(all_samples, ignore_index=True)

    # 2. Global Deduplication (Handling Model Overlaps)
    final_df['signature'] = final_df['paper_id'] + "|" + final_df['subject'] + "|" + final_df['predicate'] + "|" + final_df['object']
    initial_count = len(final_df)
    
    # Merge sources if multiple models generated the exact same triple
    final_df['source'] = final_df.groupby('signature')['source'].transform(lambda x: ', '.join(x))
    final_df = final_df.drop_duplicates(subset=['signature']).reset_index(drop=True)
    
    overlap_count = initial_count - len(final_df)
    print(f"\n   🔄 Overlaps: Found {overlap_count} triples selected by multiple models.")
    print(f"   📉 Reduced human grading workload to {len(final_df)} distinct rows.")

    # Shuffle everything so researchers don't see patterns
    final_df = final_df.sample(frac=1, random_state=CONFIG["RANDOM_SEED"]).reset_index(drop=True)

    # 3. Build the Excel Workbook
    print("\n--> Building Excel Workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Manual Evaluation"

    researcher_cols = [f"Researcher_{i}" for i in range(1, CONFIG["NUM_RESEARCHERS"] + 1)]
    headers = [
        'paper_id', 'source', 'abstract_text', 'Formatted_Triple', 
        'subject', 'predicate', 'object'
    ] + researcher_cols
    
    ws.append(headers)

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(left=Side(style='thin', color='D0D3D4'), right=Side(style='thin', color='D0D3D4'),
                         top=Side(style='thin', color='D0D3D4'), bottom=Side(style='thin', color='D0D3D4'))

    # Populate Data
    for row_idx, row in final_df.iterrows():
        excel_row = row_idx + 2
        
        ws.cell(row=excel_row, column=1, value=row.get('paper_id'))
        ws.cell(row=excel_row, column=2, value=row.get('source'))
        
        abstract_cell = ws.cell(row=excel_row, column=3)
        abstract_cell.value = apply_bold_to_entities(row.get('abstract_text'), row.get('subject'), row.get('object'))
        abstract_cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        triple_str = f"[{row.get('subject_type', 'UNKNOWN')}] {row.get('subject')} --({row.get('predicate')})--> [{row.get('object_type', 'UNKNOWN')}] {row.get('object')}"
        ws.cell(row=excel_row, column=4, value=triple_str).alignment = Alignment(vertical="top")
        
        ws.cell(row=excel_row, column=5, value=row.get('subject')).alignment = Alignment(vertical="top")
        ws.cell(row=excel_row, column=6, value=row.get('predicate')).alignment = Alignment(vertical="top")
        ws.cell(row=excel_row, column=7, value=row.get('object')).alignment = Alignment(vertical="top")

        # Row zebra striping
        row_color = "FFFFFF" if excel_row % 2 != 0 else "F8F9F9"
        fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=excel_row, column=col)
            c.fill = fill
            c.border = thin_border

    # Adjust Widths & Dropdowns
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv)
    
    start_col_letter = 'H' 
    end_col_letter = chr(ord('H') + CONFIG["NUM_RESEARCHERS"] - 1)
    for col_letter in [chr(i) for i in range(ord(start_col_letter), ord(end_col_letter) + 1)]:
        ws.column_dimensions[col_letter].width = 15
        
    dv.add(f"{start_col_letter}2:{end_col_letter}{len(final_df)+1}")
    ws.freeze_panes = "D2"

    os.makedirs(os.path.dirname(CONFIG["OUTPUT_EXCEL"]), exist_ok=True)
    wb.save(CONFIG["OUTPUT_EXCEL"])
    
    print(f"\n✅ Success! Saved {len(final_df)} optimized triples to {CONFIG['OUTPUT_EXCEL']}")

if __name__ == "__main__":
    load_and_sample()